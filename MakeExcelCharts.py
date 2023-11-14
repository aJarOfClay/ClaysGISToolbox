import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.cell.cell import Cell

# make charts across several sheets of an Excel document referencing the same columns
def chart_for_each_sheet(filename, list_of_sheets, y_axis_columns, x_axis_label_column='A', split_axes=False):
    # open workbook and for each sheet add a chart
    y_cols_int = []
    for col in y_axis_columns:
        y_cols_int.append(ord(col.lower()[0]) - 96)
    wb = load_workbook(filename)
    for sheet in list_of_sheets:
        ws = wb[sheet]
        rows_max = find_max_row(ws, 'A')
        if split_axes and len(y_axis_columns) == 2:
            chart0 = LineChart()
            data0 = Reference(ws, min_col=y_cols_int[0], max_col=y_cols_int[0], min_row=1, max_row=rows_max)
            chart0.add_data(data0, titles_from_data=True)
            j = y_axis_columns[0] + "1"
            chart0.y_axis.title = ws[j].internal_value # Cell(ws, row=1, column=y_cols_int[0])
            title = "Profile " + str(sheet)
            chart0.title = title
            chart0.x_axis.crosses = 'min'
            chart0.height = 12
            chart0.width = 25
            chart1 = LineChart()
            data1 = Reference(ws, min_col=y_cols_int[1], max_col=y_cols_int[1], min_row=1, max_row=rows_max)
            chart1.add_data(data1, titles_from_data=True)
            j = y_axis_columns[1] + "1"
            chart1.y_axis.title = ws[j].internal_value
            chart1.y_axis.axId = 200
            chart1.y_axis.crosses = 'max'
            chart1.y_axis.majorGridlines = None
            chart0 += chart1
            ws.add_chart(chart0, "C3")
        else:
            chart = LineChart()
            for col in y_cols_int:
                d = Reference(ws, min_col=col, max_col=col, min_row=1, max_row=rows_max)
                chart.add_data(d, titles_from_data=True)
            title = "Profile " + str(sheet)
            chart.title = title
            ws.add_chart(chart)

    wb.save(filename)


        # use x_axis_column and y_axis_columns with the range [2:rows_max] to create a chart in each sheet, titles come from row 1



    return

def add_single_chart():
    return

# return the number of non-empty rows in a column
def find_max_row(ws, column):
    # sanitize column input in case a letter was passed
    try:
        col_to_pass = int(column)
    except ValueError:
        col_to_pass = ord(column.lower()[0]) - 96

    num_rows = 1
    # check every hundred then work backwards
    while True:
        if ws.cell(num_rows, col_to_pass).value is not None:
            num_rows += 100
        else:
            # catch for an empty first row
            if ws.cell(num_rows+1, col_to_pass).value is not None:
                num_rows += 100
            else:
                break
    while True:
        if ws.cell(num_rows, col_to_pass).value is None:
            num_rows -= 1
        else:
            break

    return num_rows

if __name__ == "__main__":
    chart_for_each_sheet('F:\\GIS\\MartianGlaciers\\MeasuringKaseiValles\\test3.xlsx', ['A','B','C'], ['E','F'], 'C', True)
